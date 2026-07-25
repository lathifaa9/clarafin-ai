import os
import csv
import pdfplumber
import openpyxl
from typing import List, Dict, Any

def parse_document(file_path: str, filename: str) -> Dict[str, Any]:
    """
    Parses PDF, CSV, or XLSX files.
    Returns a dictionary:
    {
        "filename": str,
        "doc_type": "PDF" | "CSV" | "XLSX" | "UNKNOWN",
        "chunks": [
            {
                "text": str,
                "location": str, # e.g. "Page 1", "Row 15", "Sheet1, Cell B3"
                "page": int,     # 1-indexed page or row number
                "detail": str    # raw line or row representation
            }
        ]
    }
    """
    ext = os.path.splitext(filename)[1].lower()
    
    if ext == '.pdf':
        return parse_pdf(file_path, filename)
    elif ext == '.csv':
        return parse_csv(file_path, filename)
    elif ext in ['.xlsx', '.xls']:
        return parse_xlsx(file_path, filename)
    else:
        return {
            "filename": filename,
            "doc_type": "UNKNOWN",
            "chunks": [{"text": f"Unsupported file type: {filename}", "location": "File Header", "page": 1, "detail": ""}]
        }

def parse_pdf(file_path: str, filename: str) -> Dict[str, Any]:
    chunks = []
    has_text = False
    
    try:
        with pdfplumber.open(file_path) as pdf:
            for idx, page in enumerate(pdf.pages):
                page_num = idx + 1
                text = page.extract_text()
                
                if text and text.strip():
                    has_text = True
                    lines = text.split('\n')
                    for line_idx, line in enumerate(lines):
                        if line.strip():
                            chunks.append({
                                "text": line.strip(),
                                "location": f"Page {page_num}, Line {line_idx + 1}",
                                "page": page_num,
                                "detail": line
                            })
                else:
                    # Page has no extractable text. Fallback to OCR.
                    # We can use pytesseract if available.
                    ocr_text = run_ocr_on_page(page)
                    if ocr_text.strip():
                        has_text = True
                        lines = ocr_text.split('\n')
                        for line_idx, line in enumerate(lines):
                            if line.strip():
                                chunks.append({
                                    "text": line.strip(),
                                    "location": f"Page {page_num} (OCR), Line {line_idx + 1}",
                                    "page": page_num,
                                    "detail": line
                                })
                                
        if not has_text:
            chunks.append({
                "text": "Scanned PDF page detected, but no readable text could be extracted via OCR.",
                "location": "Page 1",
                "page": 1,
                "detail": ""
            })
            
    except Exception as e:
        chunks.append({
            "text": f"Error parsing PDF: {str(e)}",
            "location": "File Reader",
            "page": 1,
            "detail": ""
        })
        
    return {
        "filename": filename,
        "doc_type": "PDF",
        "chunks": chunks
    }

def run_ocr_on_page(pdf_page) -> str:
    """Fallback OCR using pytesseract if installed."""
    try:
        import pytesseract
        # Convert PDF page to PIL Image
        img = pdf_page.to_image(resolution=150).original
        return pytesseract.image_to_string(img)
    except Exception:
        try:
            # Try easyocr fallback
            import easyocr
            import numpy as np
            img = pdf_page.to_image(resolution=150).original
            reader = easyocr.Reader(['en'], gpu=False)
            result = reader.readtext(np.array(img), detail=0)
            return "\n".join(result)
        except Exception:
            return ""

def parse_csv(file_path: str, filename: str) -> Dict[str, Any]:
    chunks = []
    try:
        # Check if we can parse with pandas
        try:
            import pandas as pd
            df = pd.read_csv(file_path)
            for idx, row in df.iterrows():
                row_str = ", ".join([f"{col}: {val}" for col, val in row.items() if pd.notna(val)])
                chunks.append({
                    "text": row_str,
                    "location": f"Row {idx + 2}", # +2 for 1-based index and header row
                    "page": idx + 2,
                    "detail": str(row.to_dict())
                })
        except ImportError:
            # Fallback to standard csv module
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                for idx, row in enumerate(reader):
                    if header:
                        row_str = ", ".join([f"{h}: {v}" for h, v in zip(header, row)])
                    else:
                        row_str = ", ".join(row)
                    chunks.append({
                        "text": row_str,
                        "location": f"Row {idx + 2}",
                        "page": idx + 2,
                        "detail": str(row)
                    })
    except Exception as e:
        chunks.append({
            "text": f"Error parsing CSV: {str(e)}",
            "location": "File Reader",
            "page": 1,
            "detail": ""
        })
        
    return {
        "filename": filename,
        "doc_type": "CSV",
        "chunks": chunks
    }

def parse_xlsx(file_path: str, filename: str) -> Dict[str, Any]:
    chunks = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                row_values = [str(val) for val in row if val is not None]
                if row_values:
                    row_str = ", ".join(row_values)
                    chunks.append({
                        "text": f"Sheet: {sheet_name} | {row_str}",
                        "location": f"Sheet '{sheet_name}', Row {row_idx}",
                        "page": row_idx,
                        "detail": row_str
                    })
    except Exception as e:
        chunks.append({
            "text": f"Error parsing Excel file: {str(e)}",
            "location": "File Reader",
            "page": 1,
            "detail": ""
        })
        
    return {
        "filename": filename,
        "doc_type": "XLSX",
        "chunks": chunks
    }
